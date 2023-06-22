import pandas as pd
from pathlib import Path
import os
from os import remove
import sys
from datetime import datetime
import psutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from tkinter import Tk, Label, Button, messagebox, filedialog

# Funcion para cambiar el mensaje de como va el proceso
def procesos(Mensaje):
    etiquetaMensajes.config(text=Mensaje)

# Funcion para cambiar la ubicacion del mensaje de como va el proceso dependiando el texto
# Este se realizo al tanteo manualmente por que conocemos el texto q va a salir  
def ajustesUbicacionEtiqueta(xx,yy):
    etiquetaMensajes.place(x=xx,y=yy)

################################################################################################################
#    INTERFACE
ventan = Tk()
ventan.config(bg='black')
ventan.geometry('600x400') #--> ANCHO Y ALTO
ventan.resizable(0,0)#--> Impide que modifiquen el tamaño de la interface
ventan.minsize(width=600, height=400)
ventan.title('Extraer lista de mayor ') #--> titulo de la interface


# coloque un pie de pagina para colocar mi firma personal
etiqueta=Label(ventan,text="By: CAC",font=("Cascadia Code",13),fg="red",background="black")
etiqueta.place(x=0,y=370) #--> indico la ubicacion de la etiqueta dentro de la interface

#  Obtenemos el largo y  ancho de la pantalla
Ancho_de_pantalla = ventan.winfo_screenwidth() #--> Sacamos en una variable el ancho de la pantalla fisica
Alto_de_pantalla = ventan.winfo_screenheight() #--> Sacamos en una variable el alto de la pantalla fisica

ANCventana = 600 #--> Sacamos en una variable el ancho de la interface
ALTventana = 400 #--> Sacamos en una variable el alto de la interface

P_ancho = round(Ancho_de_pantalla/2-ANCventana/2)#--> Con esta operacion matematica sacamos 
                                                 #    la mitad del ancho de la pantalla.
P_alto = round(Alto_de_pantalla/2-ALTventana/2)  #--> Con esta operacion matematica sacamos 
                                                 #    la mitad del alto de la pantalla.

# Operacion realizada para centrar la interface en la pantalla 
ventan.geometry(str(ANCventana)+"x"+str(ALTventana)+"+"+str(P_ancho)+"+"+str(P_alto))

# Anuncio de como va el proceso
etiquetaMensajes=Label(ventan,text="",font=("Cascadia Code",13),fg="red",background="black")

#---------------------------------------------------------------------------------------------------------------
# ETL del libro a transformar
#---------------------------------------------------------------------------------------------------------------
def extraerLibro(): 
    ajustesUbicacionEtiqueta(200,230)
    procesos("Cargando el archivo")
    
    archivo = filedialog.askopenfilename(initialdir=str(Path.home() / "Documents"), #r"C:/", 
                                                title='Seleccione Archivo', 
                                                filetype=(('Archivos de Excel', '*.xlsx'),
                                                         ('Todos los archivos', '*.*')))
    
    
    if archivo:
        realizarETL(archivo)
    else:
        respuesta=messagebox.askyesno(message="Debe seleccionar la lista que quiere modificar. ¿Desea continuar?", title="ERROR")
        if respuesta:
            extraerLibro()
        else:
            ventan.destroy()

botonBuscar = Button(ventan, text="Buscar", font=("Cascadia Code", 13),command=extraerLibro)
botonBuscar.place(x=270,y=170) #--> Se le colocan coordenadas de ubicacion
#botonBuscar.grid(row=0,column=2)#--> Se le colocan por posicion en la pantalla repartidos entre columnas y filas
#botonBuscarr.pack()#--> cuando no se le coloca ubicacion se le coloca el pack()

def realizarETL(libroExcel):
        ajustesUbicacionEtiqueta(80,230)
        procesos("Ha iniciado el proceso ETL, paso 1 de 4")
        archivoexcel = r'{}'.format(libroExcel)
        dfListaRaw = pd.read_excel(archivoexcel)
        #messagebox.showinfo(message="Se comienza a realizar el ETL", title="ETL, paso 1")
        columnas=list(dfListaRaw.columns)
        fecha_hora_actual = datetime.now().date()
        fecha_actual = fecha_hora_actual.strftime("%d-%m-%Y")#--> Cambiar el formato de las fechas
        fecha_actual_titulo = fecha_hora_actual.strftime("%d/%m/%Y")#--> Cambiar el formato de las fechas

        Secciones=list(dfListaRaw["Linea"].unique())

        df_vacioGen = pd.DataFrame(columns=columnas)
        df_vacioMP = pd.DataFrame(columns=columnas)
        df_vacioNA = pd.DataFrame(columns=columnas)
        df_vacioIMP = pd.DataFrame(columns=columnas)

        for seccion in Secciones:
            df_temp = pd.DataFrame({'Linea': [None],
                                    'Codigo': [None],
                                    'Referencia': [None],
                                    'Ref_Original': [None],
                                    'Descripcion': [seccion],
                                    'Modelo': [None],
                                    'UM': [None],
                                    'MAR': [None],
                                    'PR': [None],
                                    'Aplicacion': [None],
                                    'Precio': [None],
                                    'EXT': [None]})
                    
            df_vacioGen = pd.concat([df_vacioGen, df_temp], ignore_index=True)
            df_vacioMP = pd.concat([df_vacioMP, df_temp], ignore_index=True)
            df_vacioNA = pd.concat([df_vacioNA, df_temp], ignore_index=True)
            df_vacioIMP = pd.concat([df_vacioIMP, df_temp], ignore_index=True)

        #   ESTAS SON LAS CONSULTAS PARA GENERAL DE AK SACO LAS DE MP, NA, IMP
            dfListaRaws=dfListaRaw[dfListaRaw['Linea'] == seccion].copy()
            dfListaRaws["PR"].fillna("NA",inplace=True)
            dfListaRaws.sort_values('Descripcion',ascending=True,inplace=True)

            listGen=dfListaRaws.copy()#-->LISTA GENERAL SIN FILTROS

            listMP=dfListaRaws.copy() 
            listMP=listMP[(listMP["PR"]=="+MP") | 
                        (listMP["PR"]=="MP")  | 
                        (listMP["PR"]=="MPCN")]#-->FILTRAMOS POR MARCA PROPIA

            listNA=dfListaRaws.copy() 
            listNA=listNA[(listNA["PR"]=="+HM") | 
                        (listNA["PR"]=="NA") ] #-->FILTRAMOS POR MARCA NACIONAL


            listIMP=dfListaRaws.copy()
            listIMP=listIMP[(listIMP["PR"]!="NA")&
                            (listIMP["PR"]!="+HM")&
                            (listIMP["PR"]!="+MP")&
                            (listIMP["PR"]!="MP")&
                            (listIMP["PR"]!="MPCN")]#-->FILTRAMOS POR MARCA IMPORTADO

            df_vacioGen = pd.concat([df_vacioGen, listGen], axis=0)
            df_vacioMP = pd.concat([df_vacioMP, listMP], axis=0)
            df_vacioNA = pd.concat([df_vacioNA, listNA], axis=0)
            df_vacioIMP = pd.concat([df_vacioIMP, listIMP], axis=0)
                    
        preListaGen=df_vacioGen[["Descripcion","Modelo","Codigo","Referencia","Ref_Original","UM", "MAR", "PR","Aplicacion","EXT","Precio"]].copy()
        preListaMP=df_vacioMP[["Descripcion","Modelo","Codigo","Referencia","Ref_Original","UM", "MAR", "PR","Aplicacion","EXT","Precio"]].copy()
        preListaNA=df_vacioNA[["Descripcion","Modelo","Codigo","Referencia","Ref_Original","UM", "MAR", "PR","Aplicacion","EXT","Precio"]].copy()
        preListaIMP=df_vacioIMP[["Descripcion","Modelo","Codigo","Referencia","Ref_Original","UM", "MAR", "PR","Aplicacion","EXT","Precio"]].copy()
        ajustesUbicacionEtiqueta(38,230)
        procesos("Se inicia la validación de los archivos, paso 2 de 4")
        #messagebox.showinfo(message="Se termina el ETL", title="ETL, paso 1")
        if os.path.exists(f"Lista de precios general de Obyco de {fecha_actual}.xlsx"):
            respuesta=messagebox.askyesno(message="Hoy, ya fue creado un libro, si está abierto será cerrado. Luego va hacer eleminado para ser reemplazado por este. ¿Desea continuar?", title="Título") 
            if respuesta:
                def close_excel_process(NomArchivo):
                    for exce in psutil.process_iter():
                        try:
                            if exce.name() == "EXCEL.EXE":
                                for handle in exce.open_files():
                                    if NomArchivo in handle.path:
                                        exce.kill()
                                        break
                        except psutil.AccessDenied:
                            pass
                close_excel_process(f"Lista de precios general de Obyco de {fecha_actual}.xlsx")
                remove(f"Lista de precios general de Obyco de {fecha_actual}.xlsx")
                ajustesUbicacionEtiqueta(80,230)
                procesos("Escribiendo el libro de Excel, paso 3 de 4")
                with pd.ExcelWriter(f"Lista de precios general de Obyco de {fecha_actual}.xlsx") as writer:
                    preListaGen.to_excel(writer, sheet_name='GENERAL', index=False)  # Guardar en la hoja 'General'
                    preListaMP.to_excel(writer, sheet_name='MARCA PROPIA', index=False)  # Guardar en la hoja 'Impo'
                    preListaNA.to_excel(writer, sheet_name='NACIONAL', index=False)  # Guardar en la hoja 'General'
                    preListaIMP.to_excel(writer, sheet_name='IMPORTADO', index=False)  # Guardar en la hoja 'Impo'

                    #messagebox.showinfo(message="Se comienza a realizar el diseño del libro", title="Diseño del Excel paso 2")
                    ajustesUbicacionEtiqueta(135,230)
                    procesos("Se comienza el diseño, paso 4 de 4")
                    hojas=["GENERAL","MARCA PROPIA","NACIONAL","IMPORTADO"]

                    for hoja in hojas:
                        book = writer.book #--> # Obtener el libro de trabajo de Excel
                        sheet_general = book[hoja] #-->Obtener la hoja 'General' del libro de trabajo

                        headers = sheet_general[1]

                        for header in headers:
                            header.alignment = Alignment(horizontal='center')

                        # Obtén el rango de celdas
                        rango_celdas = f'A2:{get_column_letter(sheet_general.max_column)}{sheet_general.max_row}'

                        # Agrega los filtros al rango de celdas
                        sheet_general.auto_filter.ref = rango_celdas
                #---------------------------------------------------------------------------------------------------
                        #Inserta una fila antes del encabezado   
                        sheet_general.insert_rows(1)

                        # Combina las celdas B1 a K1
                        sheet_general.merge_cells('B1:I1')

                        # Centrar la celda horizontalmente y verticalmente
                        sheet_general['B1'].alignment = Alignment(horizontal='center', vertical='center')

                        # Colocar negrita a la letra
                        #sheet_general['B1'].font = Font(bold=True)

                        # Aumentar el tamaño de la letra y Colocar negrita a la letra
                        sheet_general['B1'].font = Font(bold=True, size=22)
                        
                        # Combina las celdas J1 a K1
                        sheet_general.merge_cells('J1:K1')

                        # Ajustar el alto de la fila 1
                        sheet_general.row_dimensions[1].height = 63

                        # Escribe un valor en la celda combinada
                        sheet_general['B1'].value = f' LISTA DE PRECIO MAYOR DE OBYCO  {fecha_actual_titulo} \n{hoja}'
                        
                        # Ajustar texto
                        sheet_general['B1'].alignment = Alignment(wrap_text=True, horizontal='center')
                    
        #############--------------------------------------------------------------------------------------------------
                        #Insertar el logo de la empresa 
                        # imagen = Image('DatosEstrategicos.png')
                        # sheet_general.add_image(imagen, 'A1')
                        
                        #Encabezados de la tabla
                        # Obtener el rango de celdas A2:K2 y aplicar estilo de fondo azul con letras blancas
                        rango_celdas = sheet_general['A2:K2']
                        for fila in rango_celdas:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                                celda.font = Font(bold=True)
                                celda.font = Font(color="FFFFFF")
                                

                        # Nombres de la tabla informacion de la empresa 
                        # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                        rango_celdas = sheet_general['A1:K1']
                        for fila in rango_celdas:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                        
                        # FORMATO A LA PRIMERA COLUMNA
                        # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                        for fila in sheet_general[F"A3:A{sheet_general.max_row}"]:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                                celda.font = Font(bold=True)
                    
                    # Para colocar negrilla a las columnas que lo requieren
                        n = sheet_general.max_row
                        columnas = ["B", "D", "F", "H", "J", "K"]
                        for col in columnas:
                            if col=="D":
                                for fila in sheet_general[F"{col}3:{col}{sheet_general.max_row}"]:
                                        for celda in fila:
                                            celda.font = Font(bold=True)
                                            celda.alignment = Alignment(horizontal='right')
                            else:
                                for fila in sheet_general[F"{col}3:{col}{sheet_general.max_row}"]:
                                            for celda in fila:
                                                celda.font = Font(bold=True)

        ############################################################################################################
                    # Codigo para colorear las lineas donde se encuentra el nombre de las secciones 
                        lista=[]  
                        for celda in sheet_general['A']:
                            for seccion in Secciones:
                                if celda.value == seccion:
                                    fila = celda.row                           
                                    lista.append(fila)   

                        #rangoCell = sheet_general[f"A{fila}:K{sheet_general.max_column}"]
                        for i in lista:
                            for filaa in sheet_general[f"B{i}:K{i}"]:
                                for celda in filaa:
                                    celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ###############################################################################################################                                           

                # Se le aplilca formato de borde a toda la tabla 
                        borde = Border(
                            left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))

                        # Aplicar el estilo de borde a todas las celdas del rango
                        for fila in sheet_general[F"A1:K{sheet_general.max_row}"]:
                            for celda in fila:
                                celda.border = borde                  

                    # Ajusta el ancho de la columna B
                        sheet_general.column_dimensions['A'].width = 30
                        sheet_general.column_dimensions['B'].width = 20
                        sheet_general.column_dimensions['C'].width = 9
                        sheet_general.column_dimensions['D'].width = 18
                        sheet_general.column_dimensions['E'].width = 14
                        sheet_general.column_dimensions['F'].width = 4.8
                        sheet_general.column_dimensions['G'].width = 4.8
                        sheet_general.column_dimensions['H'].width = 4.5
                        sheet_general.column_dimensions['I'].width = 35.8
                        sheet_general.column_dimensions['J'].width = 6
                        sheet_general.column_dimensions['K'].width = 9.14       
                messagebox.showinfo(message="Trabajo finalizado", title="Actualizado")
                ventan.destroy()
        #---------------------------------------------------------------------------------------------------------------
        # Creacion del libro de excel y diseño de excel si no existe en la ruta
        #---------------------------------------------------------------------------------------------------------------
            else:
                ventan.destroy()

        else:
                ajustesUbicacionEtiqueta(80,230)
                procesos("Escribiendo el libro de Excel, paso 3 de 4")
                with pd.ExcelWriter(f"Lista de precios general de Obyco de {fecha_actual}.xlsx") as writer:
                    preListaGen.to_excel(writer, sheet_name='GENERAL', index=False)  # Guardar en la hoja 'General'
                    preListaMP.to_excel(writer, sheet_name='MARCA PROPIA', index=False)  # Guardar en la hoja 'Impo'
                    preListaNA.to_excel(writer, sheet_name='NACIONAL', index=False)  # Guardar en la hoja 'General'
                    preListaIMP.to_excel(writer, sheet_name='IMPORTADO', index=False)  # Guardar en la hoja 'Impo'

                    #messagebox.showinfo(message="Se comienza a realizar el diseño del libro", title="Diseño del Excel paso 2")
                    ajustesUbicacionEtiqueta(135,230)
                    procesos("Se comienza el diseño, paso 4 de 4")
                    hojas=["GENERAL","MARCA PROPIA","NACIONAL","IMPORTADO"]

                    for hoja in hojas:
                        book = writer.book #--> # Obtener el libro de trabajo de Excel
                        sheet_general = book[hoja] #-->Obtener la hoja 'General' del libro de trabajo

                        headers = sheet_general[1]

                        for header in headers:
                            header.alignment = Alignment(horizontal='center')

                        # Obtén el rango de celdas
                        rango_celdas = f'A2:{get_column_letter(sheet_general.max_column)}{sheet_general.max_row}'

                        # Agrega los filtros al rango de celdas
                        sheet_general.auto_filter.ref = rango_celdas
                #---------------------------------------------------------------------------------------------------
                        #Inserta una fila antes del encabezado   
                        sheet_general.insert_rows(1)

                        # Combina las celdas B1 a K1
                        sheet_general.merge_cells('B1:I1')

                        # Centrar la celda horizontalmente y verticalmente
                        sheet_general['B1'].alignment = Alignment(horizontal='center', vertical='center')

                        # Colocar negrita a la letra
                        #sheet_general['B1'].font = Font(bold=True)

                        # Aumentar el tamaño de la letra y Colocar negrita a la letra
                        sheet_general['B1'].font = Font(bold=True, size=22)
                        
                        # Combina las celdas J1 a K1
                        sheet_general.merge_cells('J1:K1')

                        # Ajustar el alto de la fila 1
                        sheet_general.row_dimensions[1].height = 63

                        # Escribe un valor en la celda combinada
                        sheet_general['B1'].value = f' LISTA DE PRECIO MAYOR DE OBYCO  {fecha_actual_titulo} \n{hoja}'
                        
                        # Ajustar texto
                        sheet_general['B1'].alignment = Alignment(wrap_text=True, horizontal='center')
                    
        #############--------------------------------------------------------------------------------------------------
                        #Insertar el logo de la empresa 
                        # imagen = Image('DatosEstrategicos.png')
                        # sheet_general.add_image(imagen, 'A1')
                        
                        #Encabezados de la tabla
                        # Obtener el rango de celdas A2:K2 y aplicar estilo de fondo azul con letras blancas
                        rango_celdas = sheet_general['A2:K2']
                        for fila in rango_celdas:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                                celda.font = Font(bold=True)
                                celda.font = Font(color="FFFFFF")
                                

                        # Nombres de la tabla informacion de la empresa 
                        # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                        rango_celdas = sheet_general['A1:K1']
                        for fila in rango_celdas:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                        
                        # FORMATO A LA PRIMERA COLUMNA
                        # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                        for fila in sheet_general[F"A3:A{sheet_general.max_row}"]:
                            for celda in fila:
                                celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                                celda.font = Font(bold=True)
                    
                    # Para colocar negrilla a las columnas que lo requieren
                        n = sheet_general.max_row
                        columnas = ["B", "D", "F", "H", "J", "K"]
                        for col in columnas:
                            if col=="D":
                                for fila in sheet_general[F"{col}3:{col}{sheet_general.max_row}"]:
                                        for celda in fila:
                                            celda.font = Font(bold=True)
                                            celda.alignment = Alignment(horizontal='right')
                            else:
                                for fila in sheet_general[F"{col}3:{col}{sheet_general.max_row}"]:
                                            for celda in fila:
                                                celda.font = Font(bold=True)

        ############################################################################################################
                    # Codigo para colorear las lineas donde se encuentra el nombre de las secciones 
                        lista=[]  
                        for celda in sheet_general['A']:
                            for seccion in Secciones:
                                if celda.value == seccion:
                                    fila = celda.row                           
                                    lista.append(fila)   

                        #rangoCell = sheet_general[f"A{fila}:K{sheet_general.max_column}"]
                        for i in lista:
                            for filaa in sheet_general[f"B{i}:K{i}"]:
                                for celda in filaa:
                                    celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        ###############################################################################################################                                           

                # Se le aplilca formato de borde a toda la tabla 
                        borde = Border(
                            left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))

                        # Aplicar el estilo de borde a todas las celdas del rango
                        for fila in sheet_general[F"A1:K{sheet_general.max_row}"]:
                            for celda in fila:
                                celda.border = borde                  

                    # Ajusta el ancho de la columna B
                        sheet_general.column_dimensions['A'].width = 30
                        sheet_general.column_dimensions['B'].width = 20
                        sheet_general.column_dimensions['C'].width = 9
                        sheet_general.column_dimensions['D'].width = 18
                        sheet_general.column_dimensions['E'].width = 14
                        sheet_general.column_dimensions['F'].width = 4.8
                        sheet_general.column_dimensions['G'].width = 4.8
                        sheet_general.column_dimensions['H'].width = 4.5
                        sheet_general.column_dimensions['I'].width = 35.8
                        sheet_general.column_dimensions['J'].width = 6
                        sheet_general.column_dimensions['K'].width = 9.14
                messagebox.showinfo(message="Trabajo finalizado", title="Actualizado")
                ventan.destroy()
        #---------------------------------------------------------------------------------------------------------------
        # Creacion del libro de excel y diseño de excel si no existe en la ruta
        #---------------------------------------------------------------------------------------------------------------
    
        
ventan.mainloop()
