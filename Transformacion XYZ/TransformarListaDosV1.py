import pandas as pd
from pathlib import Path
import numpy as np
import os
from os import remove
import sys
from datetime import datetime
import psutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle
from openpyxl.drawing.image import Image
from tkinter import Tk, Label, Button, Frame,  messagebox, filedialog, ttk, Scrollbar, VERTICAL, HORIZONTAL

#-----------------------------------------------------------------------------------------------------------------------------------
#Creamos la interface
#-----------------------------------------------------------------------------------------------------------------------------------

# Funcion para cambiar el mensaje de como va el proceso
def procesos(Mensaje):
    etiquetaMensajes.config(text=Mensaje)

# Funcion para cambiar la ubicacion del mensaje de como va el proceso dependiando el texto
# Este se realizo al tanteo manualmente por que conocemos el texto q va a salir  
def ajustesUbicacionEtiqueta(xx,yy):
    etiquetaMensajes.place(x=xx,y=yy)

################################################################################################################
#    INTERFACE
ventan = Tk()
ventan.config(bg='black')
ventan.geometry('600x400') #--> ANCHO Y ALTO
ventan.resizable(0,0)#--> Impide que modifiquen el tamaño de la interface
ventan.minsize(width=600, height=400)
ventan.title('Extraer lista de mayor ') #--> titulo de la interface
#ventan.iconbitmap("logo-33.ico")


# coloque un pie de pagina para colocar mi firma personal
etiqueta=Label(ventan,text="By: CAC",font=("Cascadia Code",13),fg="red",background="black")
etiqueta.place(x=0,y=370) #--> indico la ubicacion de la etiqueta dentro de la interface

#  Obtenemos el largo y  ancho de la pantalla
Ancho_de_pantalla = ventan.winfo_screenwidth() #--> Sacamos en una variable el ancho de la pantalla fisica
Alto_de_pantalla = ventan.winfo_screenheight() #--> Sacamos en una variable el alto de la pantalla fisica

ANCventana = 600 #--> Sacamos en una variable el ancho de la interface
ALTventana = 400 #--> Sacamos en una variable el alto de la interface

P_ancho = round(Ancho_de_pantalla/2-ANCventana/2)#--> Con esta operacion matematica sacamos 
                                                 #    la mitad del ancho de la pantalla.
P_alto = round(Alto_de_pantalla/2-ALTventana/2)  #--> Con esta operacion matematica sacamos 
                                                 #    la mitad del alto de la pantalla.

# Operacion realizada para centrar la interface en la pantalla 
ventan.geometry(str(ANCventana)+"x"+str(ALTventana)+"+"+str(P_ancho)+"+"+str(P_alto))

# Anuncio de como va el proceso
etiquetaMensajes=Label(ventan,text="",font=("Cascadia Code",13),fg="red",background="black")

#---------------------------------------------------------------------------------------------------------------
# Cargar lista
#---------------------------------------------------------------------------------------------------------------
def extraerLibro(): 
    ajustesUbicacionEtiqueta(200,230)
    procesos("Cargando el archivo")
    
    archivo = filedialog.askopenfilename(initialdir=str(Path.home() / "Desktop"), #r"C:/", 
                                                title='Seleccione Archivo', 
                                                filetype=(('Archivos de Excel', '*.xlsx'),
                                                         ('Todos los archivos', '*.*')))
#---------------------------------------------------------------------------------------------------------------
# Validar si cargó la lista
#---------------------------------------------------------------------------------------------------------------
    if archivo:
        realizarETL(archivo)
    else:
        respuesta=messagebox.askyesno(message="Debe seleccionar la lista que quiere modificar. ¿Desea continuar?", title="ERROR")
        if respuesta:
            extraerLibro()
        else:
            ventan.destroy()

botonBuscar = Button(ventan, text="Buscar", font=("Cascadia Code", 13),command=extraerLibro)
botonBuscar.place(x=270,y=170) #--> Se le colocan coordenadas de ubicacion

#---------------------------------------------------------------------------------------------------------------
# ETL del libro a transformar
#---------------------------------------------------------------------------------------------------------------         
def realizarETL(libroExcel):

    ajustesUbicacionEtiqueta(80,230)
    procesos("Ha iniciado el proceso ETL")
    archivoexcel = r'{}'.format(libroExcel)
    dfLista = pd.read_excel(archivoexcel,skiprows=1)
     
    fecha_hora_actual = datetime.now().date()
    fecha_actual = fecha_hora_actual.strftime("%d-%m-%Y")#--> Cambiar el formato de las fechas
    fecha_actual_titulo = fecha_hora_actual.strftime("%d/%m/%Y")#--> Cambiar el formato de las fechas
    
    #dfLista=lista[["Codigo","Referencia","Ref_Original","Etiquetas de fila","Modelo","UM","MAR","PR","Aplicacion", "Total","EXT"]].copy()
    #dfLista=lista.copy()
    procesos("ETL,Ha iniciado el cambio de nombre de columnas")
    #messagebox.showinfo("etl", "ya voy por la linea 103.")
    #-------------------------------------------------------------------------------------------------------------------------------------
    # Cambiamos los nombres de las columnas 
    #--------------------------------------------------------------------------------------------------------------------------------------
    
    dfLista.rename(columns={ 'Codigo' : 'CODIGO' },inplace=True)
    dfLista.rename(columns={ 'Referencia' : 'REFERENCIA' },inplace=True)
    dfLista.rename(columns={ 'Ref_Original' : 'REF.ORIGINAL' },inplace=True)
    dfLista.rename(columns={ 'Etiquetas de fila' : 'DESCRIPCION' },inplace=True)
    dfLista.rename(columns={ 'Modelo' : 'MODELO' },inplace=True)
    dfLista.rename(columns={ 'UM' : 'U.M' },inplace=True)
    dfLista.rename(columns={ 'MAR' : 'MAR' },inplace=True)
    dfLista.rename(columns={ 'Aplicacion' : 'APLICACIÓN' },inplace=True)
    dfLista.rename(columns={ 'Total' : 'PRECIO' },inplace=True)
    dfLista.rename(columns={ 'EXT' : 'EXT' },inplace=True)

    procesos("ETL,Ha terminado el cambio de nombre de columnas")
    #-------------------------------------------------------------------------------------------------------------------------------------
    # Arreglamos los valores NaN de la procedencia nacional ya que l sistema confunde el NA con NAN
    #-------------------------------------------------------------------------------------------------------------------------------------
    dfLista.loc[(~dfLista["CODIGO"].isna()) & (dfLista["PR"].isna()), 'PR'] ="NA"
    procesos("ETL,Ha terminado el arreglo de los valores NaN")
    #-------------------------------------------------------------------------------------------------------------------------------------
    # Cambiamos el tipo de dato de las columnas 
    #------------------------------------------------------------------------------------------------------------------------------------- 
    dfLista["CODIGO"] = dfLista["CODIGO"].astype(dtype="str")
    dfLista["PRECIO"] = dfLista["PRECIO"].round().astype(dtype="Int64")
    dfLista["EXT"] = dfLista["EXT"].round().astype(dtype="Int64")
    procesos("ETL,Ha terminado el cambio de los tipos de datos")
    #-------------------------------------------------------------------------------------------------------------------------------------
    # Cambiamos los valores de las columnas que son errados ejemplo  MPCN por MP
    #-------------------------------------------------------------------------------------------------------------------------------------
    dfLista.loc[dfLista["CODIGO"] =="nan", 'CODIGO'] =np.nan 
    dfLista.loc[dfLista["PR"] =="*MP", 'PR'] ="MP" 
    dfLista.loc[dfLista["PR"] =="MPCN", 'PR'] ="MP" 
    dfLista.loc[dfLista["PR"] =="+MP", 'PR'] ="MP"
    dfLista.loc[dfLista["PR"] =="+HM", 'PR'] ="HM" 
    dfLista.loc[dfLista["PR"] =="*AM", 'PR'] ="AM"
    dfLista.loc[dfLista["MAR"] =="GAP)", 'MAR'] ="GAP"
   
    procesos('ETL,Ha terminado el cambio de valores nulos "MPCN", "+HM" ')
    #-------------------------------------------------------------------------------------------------------------------------------------
    # Dividir la cadena en 'columna_str' en función del separador y tomar la parte izquierda
    # esto con el fin de eliminar el punto y el cero decimal que aperece cuando lo comvierto a string
    #-------------------------------------------------------------------------------------------------------------------------------------
    dfLista["CODIGO"] = dfLista['CODIGO'].str.split('.', expand=True)[0]
  

    #-------------------------------------------------------------------------------------------------------------------------------------
    #Llenamos una columna con el valor de otra, si la columna codigo tiene un valor vacio
    #--------------------------------------------------------------------------------------------------------------------------------------
    dfLista['CODIGO'].fillna(dfLista['DESCRIPCION'], inplace=True)
    

    #-------------------------------------------------------------------------------------------------------------------------------------
    # Aplicar la condición y eliminar o remaplazar con vacio el valor de 'descripcion' si es igual a 'codigo' y 'valor' está vacío
    #-------------------------------------------------------------------------------------------------------------------------------------
    dfLista.loc[(dfLista['CODIGO'] == dfLista['DESCRIPCION']) & (dfLista['MODELO'].isna()), 'DESCRIPCION'] = np.nan

    listaNueva=dfLista[["CODIGO","REFERENCIA","REF.ORIGINAL","DESCRIPCION","MODELO","U.M","MAR","PR","APLICACIÓN", "PRECIO","EXT"]].copy()
    procesos('ETL,Ha terminado el arreglo de la lista')

    # Obtener la ubicación del escritorio del usuario actual
    escritorio = os.path.expanduser("~/Desktop/exelce")
    # Especificar el nombre del archivo y la ruta completa
    procesos('Exraemos la ruta y el nombre del archivo de lista nueva')
    nombre_archivo = os.path.join(escritorio, f'Lista de precios general de Obyco de {fecha_actual}.xlsx')
    #---------------------------------------------------------------------------------------------------------------------------------------
    # Validar la existencia del archivo y que no este abierto
    #---------------------------------------------------------------------------------------------------------------------------------------
    procesos('Validamos que el archivo no exista')
    if os.path.exists(nombre_archivo):
            respuesta=messagebox.askyesno(message="Hoy, ya fue creado un libro, si está abierto, será cerrado. Luego va hacer eleminado para ser reemplazado por este. ¿Desea continuar?", title="Título") 
            if respuesta:
                def close_excel_process(NomArchivo):
                    for exce in psutil.process_iter():
                        try:
                            if exce.name() == "EXCEL.EXE":
                                for handle in exce.open_files():
                                    if NomArchivo in handle.path:
                                        exce.kill()
                                        break
                        except psutil.AccessDenied:
                            pass
                close_excel_process(f'Lista de precios general de Obyco de {fecha_actual}.xlsx')   
                remove(nombre_archivo)

    #---------------------------------------------------------------------------------------------------------------------------------------
    # Exportar el DataFrame a un archivo de Excel en el escritorio
    #---------------------------------------------------------------------------------------------------------------------------------------
                procesos('Creamos el archivo')
                with pd.ExcelWriter(nombre_archivo) as writer:
                    listaNueva.to_excel(writer, sheet_name='GENERAL', index=False)
                    book = writer.book #--> # Obtener el libro de trabajo de Excel
                    sheet_general = book["GENERAL"] #-->Obtener la hoja 'General' del libro de trabajo

                    headers = sheet_general[1]
                    for header in headers:
                        header.alignment = Alignment(horizontal='center')

                        # Obtén el rango de celdas
                        rango_celdas = f'A2:{get_column_letter(sheet_general.max_column)}{sheet_general.max_row}'
                        
                        # Agrega los filtros al rango de celdas
                        sheet_general.auto_filter.ref = rango_celdas
            #----------------------------------------------------------------------------------------------------------------------------------------------
            # Enunciado antes del los encabezados
            #----------------------------------------------------------------------------------------------------------------------------------------------      
                    #Inserta una fila antes del encabezado   
                    sheet_general.insert_rows(1)
                    # Combina las celdas B1 a K1
                    sheet_general.merge_cells('B1:I1')

                    # Centrar la celda horizontalmente y verticalmente
                    sheet_general['B1'].alignment = Alignment(horizontal='center', vertical='center')

                    # Colocar negrita a la letra
                    #sheet_general['B1'].font = Font(bold=True)

                    # Aumentar el tamaño de la letra y Colocar negrita a la letra
                    sheet_general['B1'].font = Font(bold=True, size=22)
                                    
                    # Combina las celdas J1 a K1
                    sheet_general.merge_cells('J1:K1')

                    # Ajustar el alto de la fila 1
                    sheet_general.row_dimensions[1].height = 63

                    # Escribe un valor en la celda combinada
                    sheet_general['B1'].value = f' LISTA DE PRECIO MAYOR DE OBYCO  {fecha_actual_titulo} \n{"GENERAL"}'
                                    
                    # Ajustar texto
                    sheet_general['B1'].alignment = Alignment(wrap_text=True, horizontal='center') 
                     
            #------------------------------------------------------------------------------------------------------------------------------------------------------
            # Dar estilos a los encabezados
            #------------------------------------------------------------------------------------------------------------------------------------------------------
                    #Encabezados de la tabla
                    # Obtener el rango de celdas A2:K2 y aplicar estilo de fondo azul con letras blancas
                    rango_celdas = sheet_general['A2:K2']
                    for fila in rango_celdas:
                        for celda in fila:
                            celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                            celda.font = Font(bold=True, color="FFFFFF")
                            #elda.font = Font(color="FFFFFF")
                            
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
            # Nombres de la tabla informacion de la empresa 
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
                    # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                    rango_celdas = sheet_general['A1:K1']
                    for fila in rango_celdas:
                        for celda in fila:
                            celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

            #--------------------------------------------------------------------------------------------------------------------------------------------------------
            # dar formato de numero a la columna "A"
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
                    number_format = NamedStyle(name="number_format")
                    number_format.number_format = '0'  # Ejemplo: 1234

            # Aplica el estilo a la columna deseada (por ejemplo, columna A)
                    for fila in sheet_general[F"A3:A{sheet_general.max_row}"]:
                        for celda in fila:
                            celda.style = number_format                
                            
            #########################################  CHAGP-4 ##############################################################################
            # NOSE QUE HIZO ACA
            # Configurar la alineación a la derecha 3 COLUMNA Sacada chaGP-4
            # Obtener las letras de las columnas "A" y "B" como índices numéricos
                    columnas = ["A","B","C"]
                    for col in columnas:
                        col_index = openpyxl.utils.column_index_from_string(col)

                        # Iterar sobre las filas desde la segunda hasta la última fila
                        for row in sheet_general.iter_rows(min_row=2, max_row=sheet_general.max_row, min_col=col_index, max_col=col_index):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='right')

            # Configurar la aliniacion del nombre de la columna "A", "B", "C"
                    sheet_general['A2'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet_general['B2'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet_general['C2'].alignment = Alignment(horizontal='center', vertical='center')

            ###########################################################################################################################

            # Se le aplilca formato de borde a toda la tabla 
                    borde = Border(
                                    left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                    # Aplicar el estilo de borde a todas las celdas del rango
                    for fila in sheet_general[F"A1:K{sheet_general.max_row}"]:
                        for celda in fila:
                            celda.border = borde                  

                    # Ajusta el ancho de la columna B
                    sheet_general.column_dimensions['A'].width = 8.57
                    sheet_general.column_dimensions['B'].width = 14.14
                    sheet_general.column_dimensions['C'].width = 14.71
                    sheet_general.column_dimensions['D'].width = 30
                    sheet_general.column_dimensions['E'].width = 21
                    sheet_general.column_dimensions['F'].width = 5.40
                    sheet_general.column_dimensions['G'].width = 5.83
                    sheet_general.column_dimensions['H'].width = 4.43
                    sheet_general.column_dimensions['I'].width = 27.57
                    sheet_general.column_dimensions['J'].width = 8.57
                    sheet_general.column_dimensions['K'].width = 7  
                     
                messagebox.showinfo(message="Trabajo finalizado", title="Actualizado")
                ventan.destroy()

            else:
                ventan.destroy()
    else:
            #---------------------------------------------------------------------------------------------------------------------------------------
    # Exportar el DataFrame a un archivo de Excel en el escritorio
    #---------------------------------------------------------------------------------------------------------------------------------------
                with pd.ExcelWriter(nombre_archivo) as writer:
                    listaNueva.to_excel(writer, sheet_name='GENERAL', index=False)
                    book = writer.book #--> # Obtener el libro de trabajo de Excel
                    sheet_general = book["GENERAL"] #-->Obtener la hoja 'General' del libro de trabajo

                    headers = sheet_general[1]
                    for header in headers:
                        header.alignment = Alignment(horizontal='center')

                        # Obtén el rango de celdas
                        rango_celdas = f'A2:{get_column_letter(sheet_general.max_column)}{sheet_general.max_row}'
                        
                        # Agrega los filtros al rango de celdas
                        sheet_general.auto_filter.ref = rango_celdas
            #----------------------------------------------------------------------------------------------------------------------------------------------
            # Enunciado antes del los encabezados
            #----------------------------------------------------------------------------------------------------------------------------------------------      
                    #Inserta una fila antes del encabezado   
                    sheet_general.insert_rows(1)
                    # Combina las celdas B1 a K1
                    sheet_general.merge_cells('B1:I1')

                    # Centrar la celda horizontalmente y verticalmente
                    sheet_general['B1'].alignment = Alignment(horizontal='center', vertical='center')

                    # Colocar negrita a la letra
                    #sheet_general['B1'].font = Font(bold=True)

                    # Aumentar el tamaño de la letra y Colocar negrita a la letra
                    sheet_general['B1'].font = Font(bold=True, size=22)
                                    
                    # Combina las celdas J1 a K1
                    sheet_general.merge_cells('J1:K1')

                    # Ajustar el alto de la fila 1
                    sheet_general.row_dimensions[1].height = 63

                    # Escribe un valor en la celda combinada
                    sheet_general['B1'].value = f' LISTA DE PRECIO MAYOR DE OBYCO {fecha_actual_titulo} \n{"papote GENERAL"}'
                                    
                    # Ajustar texto
                    sheet_general['B1'].alignment = Alignment(wrap_text=True, horizontal='center') 
                    
            #------------------------------------------------------------------------------------------------------------------------------------------------------
            # Dar estilos a los encabezados
            #------------------------------------------------------------------------------------------------------------------------------------------------------
                    #Encabezados de la tabla
                    # Obtener el rango de celdas A2:K2 y aplicar estilo de fondo azul con letras blancas
                    rango_celdas = sheet_general['A2:K2']
                    for fila in rango_celdas:
                        for celda in fila:
                            celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                            celda.font = Font(bold=True, color="FFFFFF")
                            #elda.font = Font(color="FFFFFF")
                            
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
            # Nombres de la tabla informacion de la empresa 
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
                    # Obtener el rango de celdas A1:K1 y aplicar estilo de fondo azul
                    rango_celdas = sheet_general['A1:K1']
                    for fila in rango_celdas:
                        for celda in fila:
                            celda.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

            #--------------------------------------------------------------------------------------------------------------------------------------------------------
            # dar formato de numero a la columna "A"
            #--------------------------------------------------------------------------------------------------------------------------------------------------------
                    number_format = NamedStyle(name="number_format")
                    number_format.number_format = '0'  # Ejemplo: 1234

            # Aplica el estilo a la columna deseada (por ejemplo, columna A)
                    for fila in sheet_general[F"A3:A{sheet_general.max_row}"]:
                        for celda in fila:
                            celda.style = number_format                
                            
            #########################################  CHAGP-4 ##############################################################################
            # NOSE QUE HIZO ACA
            # Configurar la alineación a la derecha 3 COLUMNA Sacada chaGP-4
            # Obtener las letras de las columnas "A" y "B" como índices numéricos
                    columnas = ["A","B","C"]
                    for col in columnas:
                        col_index = openpyxl.utils.column_index_from_string(col)

                        # Iterar sobre las filas desde la segunda hasta la última fila
                        for row in sheet_general.iter_rows(min_row=2, max_row=sheet_general.max_row, min_col=col_index, max_col=col_index):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='right')

            # Configurar la aliniacion del nombre de la columna "A", "B", "C"
                    sheet_general['A2'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet_general['B2'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet_general['C2'].alignment = Alignment(horizontal='center', vertical='center')

            ###########################################################################################################################

            # Se le aplilca formato de borde a toda la tabla 
                    borde = Border(
                                    left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                    # Aplicar el estilo de borde a todas las celdas del rango
                    for fila in sheet_general[F"A1:K{sheet_general.max_row}"]:
                        for celda in fila:
                            celda.border = borde                  

                    # Ajusta el ancho de la columna B
                    sheet_general.column_dimensions['A'].width = 8.57
                    sheet_general.column_dimensions['B'].width = 14.14
                    sheet_general.column_dimensions['C'].width = 14.71
                    sheet_general.column_dimensions['D'].width = 30
                    sheet_general.column_dimensions['E'].width = 21
                    sheet_general.column_dimensions['F'].width = 5.40
                    sheet_general.column_dimensions['G'].width = 5.83
                    sheet_general.column_dimensions['H'].width = 4.43
                    sheet_general.column_dimensions['I'].width = 27.57
                    sheet_general.column_dimensions['J'].width = 8.57
                    sheet_general.column_dimensions['K'].width = 7   
    
                messagebox.showinfo(message="Trabajo finalizado", title="Actualizado")
                ventan.destroy()
        #---------------------------------------------------------------------------------------------------------------
        # Creacion del libro de excel y diseño de excel si no existe en la ruta
        #---------------------------------------------------------------------------------------------------------------

ventan.mainloop()